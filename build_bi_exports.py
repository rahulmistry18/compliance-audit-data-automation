"""
Builds the Tableau Public / Power BI ready exports in data/bi_exports/ from
the audit trail produced by `python -m src.main`.

Run this after src.main so the exports reflect the latest run:
    python -m src.main
    python build_bi_exports.py
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src import config

OUT_DIR = "data/bi_exports"
os.makedirs(OUT_DIR, exist_ok=True)

HEADER_FILL = PatternFill("solid", start_color="118DFF", end_color="118DFF")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=14)
SUB_FONT = Font(name="Arial", italic=True, size=10, color="605E5C")
THIN = Side(style="thin", color="E1DFDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER


def autofit(ws, df, start_col=1):
    for i, col in enumerate(df.columns):
        width = max(len(str(col)), df[col].astype(str).map(len).max() if len(df) else 10) + 3
        ws.column_dimensions[get_column_letter(start_col + i)].width = min(width, 60)


def build_flat_export(trail: pd.DataFrame) -> pd.DataFrame:
    return pd.DataFrame({
        "Record ID": trail["entity_id"],
        "Record Type": trail["entity_type"].map({"ledger_entry": "Trade", "mandate": "Client Contract"}),
        "Check ID": trail["rule_id"],
        "Applies To": trail["domain"],
        "Result": trail["status"].map({"PASS": "Passed", "FAIL": "Had a problem"}),
        "Severity": trail["severity"].str.title(),
        "Needs Review": trail.apply(
            lambda r: "Yes" if (r["status"] == "FAIL" and r["severity"] == config.SEVERITY_HIGH) else "No",
            axis=1,
        ),
        "Finding": trail["message"],
        "Checked At (UTC)": pd.to_datetime(trail["evaluated_at"]).dt.strftime("%Y-%m-%d %H:%M"),
    })


def build_workbook(results: pd.DataFrame, path: str):
    wb = Workbook()

    ws = wb.active
    ws.title = "Results"
    ws.append(list(results.columns))
    style_header(ws, 1, len(results.columns))
    for row in results.itertuples(index=False):
        ws.append(list(row))
    for r in range(2, len(results) + 2):
        for c in range(1, len(results.columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = BORDER
            if results.columns[c - 1] == "Result":
                cell.font = Font(name="Arial", size=10, bold=True,
                                  color="107C10" if cell.value == "Passed" else "D83B01")
    autofit(ws, results)
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Rule Summary")
    ws2["A1"] = "Compliance Audit - Rule Summary"
    ws2["A1"].font = TITLE_FONT
    ws2["A2"] = "Every total below is a live formula reading the Results sheet - edit Results and these update automatically."
    ws2["A2"].font = SUB_FONT
    headers = ["Check ID", "Applies To", "Times Run", "Passed", "Had a Problem", "Needed Review"]
    ws2.append([])
    ws2.append(headers)
    style_header(ws2, 4, len(headers))

    rule_domain_pairs = (
        results[["Check ID", "Applies To"]].drop_duplicates().sort_values(["Check ID", "Applies To"]).values.tolist()
    )
    n = len(results) + 1
    start_row = 5
    for i, (rule_id, domain) in enumerate(rule_domain_pairs):
        r = start_row + i
        ws2.cell(row=r, column=1, value=rule_id)
        ws2.cell(row=r, column=2, value=domain)
        ws2.cell(row=r, column=3, value=f'=COUNTIFS(Results!$C$2:$C${n},A{r},Results!$D$2:$D${n},B{r})')
        ws2.cell(row=r, column=4, value=f'=COUNTIFS(Results!$C$2:$C${n},A{r},Results!$D$2:$D${n},B{r},Results!$E$2:$E${n},"Passed")')
        ws2.cell(row=r, column=5, value=f'=COUNTIFS(Results!$C$2:$C${n},A{r},Results!$D$2:$D${n},B{r},Results!$E$2:$E${n},"Had a problem")')
        ws2.cell(row=r, column=6, value=f'=COUNTIFS(Results!$C$2:$C${n},A{r},Results!$D$2:$D${n},B{r},Results!$G$2:$G${n},"Yes")')
        for c in range(1, 7):
            ws2.cell(row=r, column=c).font = BODY_FONT
            ws2.cell(row=r, column=c).border = BORDER

    total_row = start_row + len(rule_domain_pairs) + 1
    ws2.cell(row=total_row, column=2, value="TOTAL").font = Font(name="Arial", bold=True, size=10)
    for c, col_letter in zip(range(3, 7), ["C", "D", "E", "F"]):
        ws2.cell(row=total_row, column=c, value=f"=SUM({col_letter}{start_row}:{col_letter}{total_row-2})")
        ws2.cell(row=total_row, column=c).font = Font(name="Arial", bold=True, size=10)
        ws2.cell(row=total_row, column=c).border = BORDER

    for col, width in zip("ABCDEF", [14, 40, 12, 10, 14, 14]):
        ws2.column_dimensions[col].width = width

    ws3 = wb.create_sheet("Data Dictionary")
    ws3["A1"] = "Data Dictionary"
    ws3["A1"].font = TITLE_FONT
    dict_rows = [
        ("Record ID", "The trade or contract this check was run against (e.g. TXN-1005, MND-204)."),
        ("Record Type", '"Trade" (a ledger entry) or "Client Contract" (a mandate).'),
        ("Check ID", "Which rule was applied (see the project README for the full list)."),
        ("Applies To", 'Which region/regulator this check reflects, or "Bookkeeping"/"Governance" for internal-only checks.'),
        ("Result", '"Passed" or "Had a problem".'),
        ("Severity", "Low, Medium, or High. Only High-severity failures are flagged as needing review."),
        ("Needs Review", '"Yes" if this is a High-severity failure that should be looked at by a person.'),
        ("Finding", "Plain-text explanation of what the check found."),
        ("Checked At (UTC)", "Timestamp the check was run."),
    ]
    ws3.append([])
    ws3.append(["Column", "Meaning"])
    style_header(ws3, 3, 2)
    for i, (col, meaning) in enumerate(dict_rows):
        r = 4 + i
        ws3.cell(row=r, column=1, value=col).font = Font(name="Arial", bold=True, size=10)
        ws3.cell(row=r, column=2, value=meaning).font = BODY_FONT
        ws3.cell(row=r, column=1).border = BORDER
        ws3.cell(row=r, column=2).border = BORDER
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 90

    ws4 = wb.create_sheet("Data Source")
    ws4["A1"] = "Data Source Disclosure"
    ws4["A1"].font = TITLE_FONT
    lines = [
        "",
        "Trade and contract records: 0% real. All records in this workbook are fabricated for",
        "demonstration. No real client, trade, or company appears anywhere in this dataset.",
        "",
        "Rules and deadlines applied: 100% real. Reporting deadlines, required data fields, and",
        "regulator names reflect actual public regulatory guidance (EMIR, MiFID II, Dodd-Frank/",
        "CFTC/SEC, ASIC, HKMA/SFC, MAS), current as of mid-2026. Simplified for demonstration;",
        "not legal advice.",
        "",
        "Regenerate a fresh (equally fictional) dataset with: python generate_sample_data.py",
        "Full source: see the README in the GitHub repository.",
    ]
    for i, line in enumerate(lines):
        ws4.cell(row=2 + i, column=1, value=line).font = BODY_FONT
    ws4.column_dimensions["A"].width = 100

    wb.save(path)


if __name__ == "__main__":
    trail = pd.read_csv(config.AUDIT_TRAIL_CSV)
    results = build_flat_export(trail)
    results.to_csv(f"{OUT_DIR}/audit_results.csv", index=False)
    build_workbook(results, f"{OUT_DIR}/compliance_audit_workbook.xlsx")
    print(f"Wrote {OUT_DIR}/audit_results.csv ({len(results)} rows)")
    print(f"Wrote {OUT_DIR}/compliance_audit_workbook.xlsx")
